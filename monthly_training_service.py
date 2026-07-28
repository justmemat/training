"""Business rules and Excel reporting for monthly training sessions."""

from datetime import date
import os
from pathlib import Path, PureWindowsPath
from typing import Any, Callable
from uuid import uuid4

from training_directory_service import full_name


MONTHLY_REPORT_DIRECTORY = Path(r"T:\BAE\Training\Monthly\Reports")
MONTHLY_REPORT_NAME = "Monthly Training History.xlsx"


def create_session_record(
    *,
    presentation_date: date,
    instructor_id: str,
    attendee_ids: list[str],
    presentation_path: str,
    team_members: list[dict[str, Any]],
) -> dict[str, Any]:
    """Validate and create a monthly-training session record."""
    member_ids = {str(member.get("id", "")) for member in team_members}
    if instructor_id not in member_ids:
        raise ValueError("Select a valid instructor.")
    unique_attendees = list(dict.fromkeys(attendee_ids))
    if not unique_attendees:
        raise ValueError("Select at least one team member in attendance.")
    if any(attendee_id not in member_ids for attendee_id in unique_attendees):
        raise ValueError("Attendance contains an unknown team member.")
    selected_file = Path(presentation_path)
    if not presentation_path.strip() or not selected_file.is_file():
        raise ValueError("Select the file that was presented.")
    return {
        "id": str(uuid4()),
        "date": presentation_date.isoformat(),
        "instructor_id": instructor_id,
        "attendee_ids": unique_attendees,
        "presentation_path": str(selected_file),
        "file_name": PureWindowsPath(str(selected_file)).name,
    }


def sorted_sessions(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return monthly training sessions newest first."""
    valid_records = [record for record in records if isinstance(record, dict)]
    return sorted(
        valid_records,
        key=lambda record: str(record.get("date", "")),
        reverse=True,
    )


def open_presentation_file(
    record: dict[str, Any], opener: Callable[[str], Any] | None = None
) -> Path:
    """Open a saved presentation with the operating system's associated app."""
    presentation_path = Path(str(record.get("presentation_path", "")).strip())
    if not str(presentation_path) or not presentation_path.is_file():
        raise FileNotFoundError(
            f"The saved presentation could not be found: {presentation_path}"
        )
    if opener is None:
        startfile = getattr(os, "startfile", None)
        if startfile is None:
            raise OSError("Opening presentations is only supported by this app on Windows.")
        opener = startfile
    opener(str(presentation_path))
    return presentation_path


def generate_monthly_history_report(
    records: list[dict[str, Any]],
    team_members: list[dict[str, Any]],
    *,
    output_directory: Path = MONTHLY_REPORT_DIRECTORY,
) -> Path:
    """Create or replace a workbook with newest-year training tabs first."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    output_directory.mkdir(parents=True, exist_ok=True)
    by_id = {str(member.get("id", "")): member for member in team_members}
    dated_records: list[tuple[date, dict[str, Any]]] = []
    for record in sorted_sessions(records):
        try:
            dated_records.append((date.fromisoformat(str(record.get("date", ""))), record))
        except ValueError:
            continue
    years = sorted({training_date.year for training_date, _ in dated_records}, reverse=True)
    if not years:
        years = [date.today().year]

    workbook = Workbook()
    workbook.remove(workbook.active)
    for year in years:
        sheet = workbook.create_sheet(str(year))
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = "A3:D3"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 32
        sheet.column_dimensions["C"].width = 55
        sheet.column_dimensions["D"].width = 40
        sheet.merge_cells("A1:D1")
        sheet["A1"] = f"Monthly Training History — {year}"
        sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        sheet["A1"].fill = PatternFill("solid", fgColor="303F9F")
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append([])
        sheet.append(["Training Date", "Instructor", "Attendance", "Presented File"])
        for training_date, record in dated_records:
            if training_date.year != year:
                continue
            attendees = ", ".join(
                full_name(by_id.get(str(attendee_id))) or "Unknown"
                for attendee_id in record.get("attendee_ids", [])
            )
            sheet.append(
                [
                    training_date,
                    full_name(by_id.get(str(record.get("instructor_id", ""))))
                    or "Unknown",
                    attendees,
                    str(record.get("file_name", "")),
                ]
            )
            sheet.cell(sheet.max_row, 1).number_format = "dd mmm yyyy"
            sheet.cell(sheet.max_row, 1).alignment = Alignment(horizontal="left")
            sheet.cell(sheet.max_row, 3).alignment = Alignment(wrap_text=True)

    output_path = output_directory / MONTHLY_REPORT_NAME
    workbook.save(output_path)
    return output_path


def open_monthly_history_report(
    report_path: Path, opener: Callable[[str], Any] | None = None
) -> Path:
    """Open the generated workbook with the operating system's associated app."""
    if not report_path.is_file():
        raise FileNotFoundError(f"The monthly training report could not be found: {report_path}")
    if opener is None:
        startfile = getattr(os, "startfile", None)
        if startfile is None:
            raise OSError("Opening the Excel report is only supported by this app on Windows.")
        opener = startfile
    opener(str(report_path))
    return report_path
