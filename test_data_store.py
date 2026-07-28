"""Tests for the feature-based JSON storage layer."""

import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import Mock, patch

import data_store
from team_member_service import display_name, role_sort_key, upsert_member
from trainee_service import (
    TRAINING_PHASES,
    ensure_profile,
    format_start_date,
    update_profile,
)
from training_directory_service import (
    add_business_days,
    build_guide_fields,
    create_trainee_folders,
    trainee_directory_exists,
)
from training_report_service import build_report_fields
from history_report_service import (
    business_days_used,
    instructor_percentages,
    training_end_date,
)
from monthly_training_service import (
    MONTHLY_REPORT_DIRECTORY,
    MONTHLY_REPORT_NAME,
    create_session_record,
    generate_monthly_history_report,
    open_monthly_history_report,
    open_presentation_file,
    sorted_sessions,
)
from training_history_service import (
    create_history_record,
    open_report_file,
    report_file_uri,
    trainee_history,
)


class DataStoreTests(unittest.TestCase):
    def test_flet_is_pinned_to_compatible_version(self) -> None:
        requirements = Path("requirements.txt").read_text(encoding="utf-8")
        self.assertIn("flet==0.86.3", requirements)
        self.assertIn("flet-desktop==0.86.3", requirements)

    def test_default_data_location_uses_shared_assets_directory(self) -> None:
        self.assertEqual(
            str(data_store.DATA_DIRECTORY),
            r"T:\BAE\Training\Onboarding\Masters\App\Assets",
        )

    def test_initializes_separate_files_and_round_trips_records(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            with patch.object(data_store, "DATA_DIRECTORY", Path(directory)):
                data_store.initialize_data_files()

                self.assertEqual(
                    {path.name for path in Path(directory).iterdir()},
                    set(data_store.DATA_FILES.values()),
                )
                records = [{"name": "Example Team Member"}]
                data_store.save_records("team_members", records)
                self.assertEqual(data_store.load_records("team_members"), records)

    def test_unknown_feature_is_rejected(self) -> None:
        with self.assertRaisesRegex(ValueError, "Unknown data feature"):
            data_store.load_records("unknown")


class TeamMemberServiceTests(unittest.TestCase):
    def test_add_update_and_display_member(self) -> None:
        members: list[dict] = []
        created = upsert_member(
            members,
            first_name="  Jamie ",
            last_name="Rivera",
            operating_initials=" jr ",
            email=" jamie.rivera@example.com ",
            is_manager=True,
            is_training_lead=False,
        )
        self.assertEqual(display_name(created), "J. Rivera")
        self.assertEqual(created["operating_initials"], "JR")
        self.assertEqual(created["email"], "jamie.rivera@example.com")
        self.assertTrue(created["is_manager"])
        self.assertFalse(created["is_trainee"])

        updated = upsert_member(
            members,
            first_name="James",
            last_name="Rivera",
            operating_initials="JR",
            email="james.rivera@example.com",
            is_manager=False,
            is_training_lead=True,
            member_id=created["id"],
        )
        self.assertEqual(len(members), 1)
        self.assertEqual(updated["first_name"], "James")
        self.assertTrue(updated["is_training_lead"])

    def test_assigning_training_lead_replaces_previous_lead(self) -> None:
        members: list[dict] = []
        first = upsert_member(
            members,
            first_name="Alex",
            last_name="One",
            operating_initials="AO",
            email="",
            is_manager=False,
            is_training_lead=True,
        )
        second = upsert_member(
            members,
            first_name="Blair",
            last_name="Two",
            operating_initials="BT",
            email="",
            is_manager=False,
            is_training_lead=True,
        )
        self.assertFalse(first["is_training_lead"])
        self.assertTrue(second["is_training_lead"])

    def test_required_and_unique_operating_initials(self) -> None:
        members: list[dict] = []
        upsert_member(
            members,
            first_name="Alex",
            last_name="One",
            operating_initials="AO",
            email="",
            is_manager=False,
            is_training_lead=False,
        )
        with self.assertRaisesRegex(ValueError, "unique"):
            upsert_member(
                members,
                first_name="Another",
                last_name="Operator",
                operating_initials="ao",
                email="",
                is_manager=False,
                is_training_lead=False,
            )

    def test_optional_email_is_validated(self) -> None:
        with self.assertRaisesRegex(ValueError, "valid email"):
            upsert_member(
                [],
                first_name="Alex",
                last_name="One",
                operating_initials="AO",
                email="not-an-email",
                is_manager=False,
                is_training_lead=False,
            )

    def test_members_sort_by_role_then_name(self) -> None:
        members = [
            {"first_name": "Zoe", "last_name": "Member"},
            {"first_name": "Taylor", "last_name": "Lead", "is_training_lead": True},
            {"first_name": "Morgan", "last_name": "Boss", "is_manager": True},
            {"first_name": "Aaron", "last_name": "Able", "is_manager": True},
        ]
        ordered = sorted(members, key=role_sort_key)
        self.assertEqual(
            [member["first_name"] for member in ordered],
            ["Aaron", "Morgan", "Taylor", "Zoe"],
        )


class TraineeServiceTests(unittest.TestCase):
    def test_start_date_uses_readable_display_format(self) -> None:
        self.assertEqual(format_start_date("2026-07-27"), "27 Jul 2026")
        self.assertEqual(format_start_date(""), "")

    def test_profile_creation_and_update(self) -> None:
        profiles: list[dict] = []
        profile = ensure_profile(profiles, "member-1")
        self.assertEqual(profile["training_phase"], "Ground School")
        self.assertEqual(profile["start_date"], "")

        update_profile(
            profile,
            start_date="2026-07-27",
            training_phase="Phase Two",
            primary_instructor_id="instructor-1",
            secondary_instructor_id="instructor-2",
            manager_id="manager-1",
            team_members=[
                {"id": "instructor-1"},
                {"id": "instructor-2"},
                {"id": "manager-1", "is_manager": True},
            ],
        )
        self.assertEqual(profile["start_date"], "2026-07-27")
        self.assertEqual(profile["training_phase"], "Phase Two")
        self.assertEqual(profile["primary_instructor_id"], "instructor-1")
        self.assertEqual(profile["secondary_instructor_id"], "instructor-2")
        self.assertEqual(profile["manager_id"], "manager-1")
        self.assertIs(ensure_profile(profiles, "member-1"), profile)
        self.assertEqual(len(profiles), 1)

    def test_training_phase_and_date_are_validated(self) -> None:
        profile = ensure_profile([], "member-1")
        with self.assertRaisesRegex(ValueError, "YYYY-MM-DD"):
            update_profile(
                profile, start_date="07/27/2026", training_phase=TRAINING_PHASES[0]
            )
        with self.assertRaisesRegex(ValueError, "Training Phase"):
            update_profile(
                profile, start_date="2026-07-27", training_phase="Phase Four"
            )

    def test_assigned_manager_requires_manager_role(self) -> None:
        profile = ensure_profile([], "trainee-1")
        with self.assertRaisesRegex(ValueError, "Manager role"):
            update_profile(
                profile,
                start_date="2026-07-27",
                training_phase="Ground School",
                manager_id="member-1",
                team_members=[{"id": "member-1", "is_manager": False}],
            )

    def test_instructors_must_reference_team_members(self) -> None:
        profile = ensure_profile([], "trainee-1")
        with self.assertRaisesRegex(ValueError, "valid instructor"):
            update_profile(
                profile,
                start_date="2026-07-27",
                training_phase="Ground School",
                primary_instructor_id="missing",
                team_members=[],
            )


class TrainingDirectoryServiceTests(unittest.TestCase):
    def test_trainee_directory_contains_reports_subfolder(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            trainee_directory = create_trainee_folders(Path(directory), "JR")
            self.assertEqual(trainee_directory, Path(directory) / "JR")
            self.assertTrue(trainee_directory.is_dir())
            self.assertTrue((trainee_directory / "Reports").is_dir())
            self.assertTrue(
                trainee_directory_exists(
                    {"operating_initials": "jr"}, Path(directory)
                )
            )
            with self.assertRaisesRegex(FileExistsError, "already exists"):
                create_trainee_folders(Path(directory), "JR")

    def test_business_day_calculation_skips_weekends(self) -> None:
        self.assertEqual(add_business_days(date(2026, 7, 24), 1), date(2026, 7, 27))
        self.assertEqual(add_business_days(date(2026, 7, 27), 30), date(2026, 9, 7))

    def test_pdf_fields_are_built_from_program_data(self) -> None:
        trainee = {
            "id": "trainee",
            "first_name": "Jamie",
            "last_name": "Rivera",
            "operating_initials": "jr",
        }
        members = [
            trainee,
            {"id": "primary", "first_name": "Pat", "last_name": "Primary"},
            {"id": "secondary", "first_name": "Sam", "last_name": "Second"},
            {
                "id": "lead",
                "first_name": "Lee",
                "last_name": "Lead",
                "is_training_lead": True,
            },
            {
                "id": "manager",
                "first_name": "Morgan",
                "last_name": "Manager",
                "is_manager": True,
            },
        ]
        profile = {
            "start_date": "2026-07-27",
            "primary_instructor_id": "primary",
            "secondary_instructor_id": "secondary",
            "manager_id": "manager",
        }
        fields = build_guide_fields(trainee, profile, members)
        self.assertEqual(fields["NAME"], "Jamie Rivera")
        self.assertEqual(fields["INITIALS"], "JR")
        self.assertEqual(fields["PRIMARY"], "Pat Primary")
        self.assertEqual(fields["SECONDARY"], "Sam Second")
        self.assertEqual(fields["LEAD"], "Lee Lead")
        self.assertEqual(fields["MANAGER"], "Morgan Manager")
        self.assertEqual(fields["StartDate"], "27 Jul 2026")
        self.assertEqual(fields["CheckOne"], "07 Sep 2026")
        self.assertEqual(fields["StudentName"], "Jamie Rivera")

    def test_pypdf_is_declared_as_an_application_dependency(self) -> None:
        requirements = Path("requirements.txt").read_text(encoding="utf-8")
        self.assertIn("pypdf", requirements)


class TrainingReportServiceTests(unittest.TestCase):
    def test_daily_report_fields_are_built_from_program_and_user_data(self) -> None:
        trainee = {
            "id": "trainee",
            "first_name": "Jamie",
            "last_name": "Rivera",
            "operating_initials": "JR",
        }
        members = [
            trainee,
            {
                "id": "primary",
                "first_name": "Pat",
                "last_name": "Primary",
                "operating_initials": "PP",
            },
            {"id": "secondary", "first_name": "Sam", "last_name": "Second"},
            {
                "id": "lead",
                "first_name": "Lee",
                "last_name": "Lead",
                "operating_initials": "LL",
                "is_training_lead": True,
            },
        ]
        fields = build_report_fields(
            trainee,
            {
                "primary_instructor_id": "primary",
                "secondary_instructor_id": "secondary",
            },
            members,
            instructor_id="primary",
            training_summary="Completed lesson one.",
            instructor_comments="Good progress.",
            report_date=date(2026, 7, 27),
        )
        self.assertEqual(fields["Trainees_Name"], "Jamie Rivera")
        self.assertEqual(fields["Trainees_Initials"], "JR")
        self.assertEqual(fields["Date"], "27 Jul 2026")
        self.assertEqual(fields["Primary_Instructor"], "Pat Primary")
        self.assertEqual(fields["Secondary_Instructor"], "Sam Second")
        self.assertEqual(fields["Training_Lead"], "Lee Lead")
        self.assertEqual(fields["Training_Summary"], "Completed lesson one.")
        self.assertEqual(fields["Instructor_Comments"], "Good progress.")
        self.assertEqual(fields["Instructors_Initials"], "PP")
        self.assertNotIn("Instructor_Initials", fields)
        self.assertEqual(fields["Trainees_Initials1"], "JR")
        self.assertEqual(fields["Training_Lead1"], "LL")

    def test_report_requires_summary_comments_and_valid_instructor(self) -> None:
        trainee = {"first_name": "Jamie", "operating_initials": "JR"}
        lead = {"id": "lead", "is_training_lead": True}
        with self.assertRaisesRegex(ValueError, "valid instructor"):
            build_report_fields(
                trainee,
                {},
                [lead],
                instructor_id="missing",
                training_summary="Summary",
                instructor_comments="Comments",
            )

    def test_blank_instructor_comments_leave_pdf_field_unchanged(self) -> None:
        trainee = {"first_name": "Jamie", "operating_initials": "JR"}
        members = [
            {"id": "instructor", "operating_initials": "II"},
            {"id": "lead", "is_training_lead": True, "operating_initials": "LL"},
        ]
        fields = build_report_fields(
            trainee,
            {},
            members,
            instructor_id="instructor",
            training_summary="Required summary",
            instructor_comments="   ",
        )
        self.assertNotIn("Instructor_Comments", fields)
        self.assertEqual(fields["Training_Summary"], "Required summary")


class TrainingHistoryServiceTests(unittest.TestCase):
    def test_history_records_date_and_selected_instructor(self) -> None:
        record = create_history_record(
            trainee_id="trainee-1",
            instructor_id="instructor-1",
            report_date=date(2026, 7, 27),
            report_path=r"T:\Reports\report.pdf",
        )
        self.assertEqual(record["trainee_id"], "trainee-1")
        self.assertEqual(record["instructor_id"], "instructor-1")
        self.assertEqual(record["date"], "2026-07-27")
        self.assertEqual(record["file_name"], "report.pdf")
        self.assertEqual(record["report_path"], r"T:\Reports\report.pdf")

    def test_trainee_history_is_filtered_and_newest_first(self) -> None:
        records = [
            {"trainee_id": "one", "date": "2026-07-01"},
            {"trainee_id": "two", "date": "2026-08-01"},
            {"trainee_id": "one", "date": "2026-07-27"},
        ]
        history = trainee_history(records, "one")
        self.assertEqual([entry["date"] for entry in history], ["2026-07-27", "2026-07-01"])

    def test_report_location_converts_to_file_uri(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            report = Path(directory) / "daily report.pdf"
            report.touch()
            uri = report_file_uri({"report_path": str(report)})
            self.assertTrue(uri.startswith("file:"))
            self.assertIn("daily%20report.pdf", uri)

    def test_report_file_uses_operating_system_opener(self) -> None:
        opened: list[str] = []
        with tempfile.TemporaryDirectory() as directory:
            report = Path(directory) / "daily report.pdf"
            report.touch()
            result = open_report_file(
                {"report_path": str(report)}, opener=opened.append
            )
            self.assertEqual(result, report)
            self.assertEqual(opened, [str(report)])


class HistoryReportServiceTests(unittest.TestCase):
    def test_training_end_date_is_90_business_days_after_start(self) -> None:
        self.assertEqual(training_end_date(date(2026, 7, 27)), date(2026, 11, 30))

    def test_business_days_used_excludes_start_and_weekends(self) -> None:
        self.assertEqual(
            business_days_used(date(2026, 7, 24), date(2026, 7, 27)), 1
        )
        self.assertEqual(
            business_days_used(date(2026, 7, 27), date(2026, 7, 27)), 0
        )

    def test_instructor_percentages_use_selected_trainees_history(self) -> None:
        percentages = instructor_percentages(
            [
                {"trainee_id": "one", "instructor_id": "a", "date": "2026-07-01"},
                {"trainee_id": "one", "instructor_id": "a", "date": "2026-07-02"},
                {"trainee_id": "one", "instructor_id": "b", "date": "2026-07-03"},
                {"trainee_id": "two", "instructor_id": "b", "date": "2026-07-04"},
            ],
            "one",
        )
        self.assertAlmostEqual(percentages["a"], 2 / 3)
        self.assertAlmostEqual(percentages["b"], 1 / 3)


class MonthlyTrainingServiceTests(unittest.TestCase):
    def test_history_report_uses_single_shared_workbook_path(self) -> None:
        self.assertEqual(
            str(MONTHLY_REPORT_DIRECTORY), r"T:\BAE\Training\Monthly\Reports"
        )
        self.assertEqual(MONTHLY_REPORT_NAME, "Monthly Training History.xlsx")

    def test_history_report_has_newest_year_tabs_first(self) -> None:
        from openpyxl import load_workbook

        records = [
            {
                "date": "2025-05-01",
                "instructor_id": "one",
                "attendee_ids": ["one"],
                "file_name": "Older.pptx",
            },
            {
                "date": "2026-06-01",
                "instructor_id": "one",
                "attendee_ids": ["one", "two"],
                "file_name": "Newer.pptx",
            },
        ]
        members = [
            {"id": "one", "first_name": "Jamie", "last_name": "Rivera"},
            {"id": "two", "first_name": "Morgan", "last_name": "Lee"},
        ]
        with tempfile.TemporaryDirectory() as directory:
            report = generate_monthly_history_report(
                records, members, output_directory=Path(directory)
            )
            workbook = load_workbook(report)

        self.assertEqual(workbook.sheetnames, ["2026", "2025"])
        self.assertEqual(workbook["2026"]["D4"].value, "Newer.pptx")
        self.assertEqual(workbook["2025"]["D4"].value, "Older.pptx")

    def test_generated_history_report_uses_operating_system_opener(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            report = Path(directory) / MONTHLY_REPORT_NAME
            report.touch()
            opener = Mock()

            result = open_monthly_history_report(report, opener=opener)

            self.assertEqual(result, report)
            opener.assert_called_once_with(str(report))

    def test_session_records_file_date_instructor_and_attendance(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            presentation = Path(directory) / "presentation.pdf"
            presentation.touch()
            members = [{"id": "one"}, {"id": "two"}]
            record = create_session_record(
                presentation_date=date(2026, 7, 27),
                instructor_id="one",
                attendee_ids=["one", "two", "one"],
                presentation_path=str(presentation),
                team_members=members,
            )
            self.assertEqual(record["date"], "2026-07-27")
            self.assertEqual(record["instructor_id"], "one")
            self.assertEqual(record["attendee_ids"], ["one", "two"])
            self.assertEqual(record["presentation_path"], str(presentation))

    def test_sessions_are_sorted_newest_first(self) -> None:
        sessions = sorted_sessions(
            [{"date": "2026-07-01"}, "invalid", {"date": "2026-08-01"}]
        )
        self.assertEqual([session["date"] for session in sessions], ["2026-08-01", "2026-07-01"])

    def test_presentation_file_uses_operating_system_opener(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            presentation = Path(directory) / "lesson.pptx"
            presentation.touch()
            opener = Mock()

            result = open_presentation_file(
                {"presentation_path": str(presentation)}, opener=opener
            )

            self.assertEqual(result, presentation)
            opener.assert_called_once_with(str(presentation))

    def test_missing_presentation_file_cannot_be_opened(self) -> None:
        with self.assertRaisesRegex(FileNotFoundError, "could not be found"):
            open_presentation_file({"presentation_path": "missing.pptx"}, opener=Mock())


if __name__ == "__main__":
    unittest.main()
