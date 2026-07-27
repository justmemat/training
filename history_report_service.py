"""Generate an Excel summary of a trainee's training history."""

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from training_directory_service import (
    TRAINING_DIRECTORY_ROOT,
    add_business_days,
    full_name,
)
from training_history_service import trainee_history


def business_days_used(start: date, through: date) -> int:
    """Count weekdays used after the start date through the requested date."""
    if through <= start:
        return 0
    current = start + timedelta(days=1)
    used = 0
    while current <= through:
        if current.weekday() < 5:
            used += 1
        current += timedelta(days=1)
    return used


def training_end_date(start: date) -> date:
    """Return the end of the trainee's 90-business-day training allotment."""
    return add_business_days(start, 90)


def instructor_percentages(
    history: list[dict[str, Any]], trainee_id: str
) -> dict[str, float]:
    """Calculate each instructor's share of the trainee's recorded reports."""
    entries = trainee_history(history, trainee_id)
    total = len(entries)
    if not total:
        return {}
    counts: dict[str, int] = {}
    for entry in entries:
        instructor_id = str(entry.get("instructor_id", ""))
        counts[instructor_id] = counts.get(instructor_id, 0) + 1
    return {
        instructor_id: count / total
        for instructor_id, count in counts.items()
    }


def generate_history_report(
    trainee: dict[str, Any],
    profile: dict[str, Any],
    team_members: list[dict[str, Any]],
    history: list[dict[str, Any]],
    *,
    output_root: Path = TRAINING_DIRECTORY_ROOT,
    report_date: date | None = None,
) -> Path:
    """Create or replace the trainee's Excel Training History Report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    initials = str(trainee.get("operating_initials", "")).strip().upper()
    reports_directory = output_root / initials / "Reports"
    if not reports_directory.is_dir():
        raise FileNotFoundError(
            "Create the trainee's training directory before generating history."
        )
    start_text = str(profile.get("start_date", ""))
    if not start_text:
        raise ValueError("Assign a trainee start date before generating history.")
    try:
        start = date.fromisoformat(start_text)
    except ValueError as error:
        raise ValueError("The trainee start date is invalid.") from error

    today = report_date or date.today()
    entries = trainee_history(history, str(trainee.get("id", "")))
    percentages = instructor_percentages(history, str(trainee.get("id", "")))
    by_id = {str(member.get("id", "")): member for member in team_members}
    lead = next(
        (member for member in team_members if member.get("is_training_lead")), None
    )
    used_days = business_days_used(start, today)
    end_date = training_end_date(start)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training History"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 38
    sheet.column_dimensions["C"].width = 24
    title_fill = PatternFill("solid", fgColor="303F9F")
    section_fill = PatternFill("solid", fgColor="C5CAE9")

    sheet.merge_cells("A1:C1")
    sheet["A1"] = f"Training History Report - {full_name(trainee)} ({initials})"
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet["A3"] = "Start Date"
    sheet["B3"] = start
    sheet["B3"].number_format = "dd mmm yyyy"
    sheet["A4"] = "End Date"
    sheet["B4"] = end_date
    sheet["B4"].number_format = "dd mmm yyyy"
    sheet["A5"] = "Total Days Training"
    sheet["B5"] = used_days
    sheet["A6"] = "Total Training Time Used"
    sheet["B6"] = used_days / 90
    sheet["B6"].number_format = "0.0%"

    sheet.merge_cells("A8:C8")
    sheet["A8"] = "Training Team"
    sheet["A8"].font = Font(bold=True)
    sheet["A8"].fill = section_fill
    team_rows = [
        ("Primary Instructor", profile.get("primary_instructor_id", "")),
        ("Secondary Instructor", profile.get("secondary_instructor_id", "")),
        ("Assigned Manager", profile.get("manager_id", "")),
    ]
    for row, (role, member_id) in enumerate(team_rows, start=9):
        sheet.cell(row, 1, role)
        sheet.cell(row, 2, full_name(by_id.get(str(member_id))) or "Unassigned")
    sheet["A12"] = "Training Lead"
    sheet["B12"] = full_name(lead) or "Unassigned"

    sheet.merge_cells("A14:C14")
    sheet["A14"] = "Instructor Share of Recorded Training"
    sheet["A14"].font = Font(bold=True)
    sheet["A14"].fill = section_fill
    sheet.append(["Instructor", "Reports", "Percentage"])
    percentage_row = 16
    for instructor_id, percentage in sorted(
        percentages.items(),
        key=lambda item: full_name(by_id.get(item[0])).lower(),
    ):
        count = sum(
            entry.get("instructor_id") == instructor_id for entry in entries
        )
        sheet.cell(percentage_row, 1, full_name(by_id.get(instructor_id)) or "Unknown")
        sheet.cell(percentage_row, 2, count)
        sheet.cell(percentage_row, 3, percentage).number_format = "0.0%"
        percentage_row += 1

    history_start = percentage_row + 1
    sheet.merge_cells(start_row=history_start, start_column=1, end_row=history_start, end_column=3)
    sheet.cell(history_start, 1, "Days Trained and Instructor").fill = section_fill
    sheet.cell(history_start, 1).font = Font(bold=True)
    sheet.append(["Training Date", "Instructor"])
    for entry in entries:
        entry_date = date.fromisoformat(str(entry.get("date", "")))
        sheet.append(
            [
                entry_date,
                full_name(by_id.get(str(entry.get("instructor_id", "")))) or "Unknown",
            ]
        )
        sheet.cell(sheet.max_row, 1).number_format = "dd mmm yyyy"
        sheet.cell(sheet.max_row, 1).alignment = Alignment(horizontal="left")

    output_path = reports_directory / f"Training History Report - {initials}.xlsx"
    workbook.save(output_path)
    return output_path
